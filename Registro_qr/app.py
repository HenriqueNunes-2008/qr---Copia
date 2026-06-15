from flask import Flask, request, render_template, jsonify, Response, redirect, url_for, session, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, DoughnutChart
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import json
from datetime import datetime, date, time, timedelta
from flask_sqlalchemy import SQLAlchemy
from dotenv import load_dotenv
from flask_bcrypt import Bcrypt # Para hash de senhas

# Carrega as variáveis de ambiente do arquivo .env antes de qualquer configuração
load_dotenv()

app = Flask(__name__)

# Nome do arquivo Excel para registros e gráficos
EXCEL_FILE = "registros.xlsx"

# --- Configurações do Banco de Dados PostgreSQL com SQLAlchemy ---
# A URL do banco de dados será lida de uma variável de ambiente.
# O uso de aspas no .env e %23 para o '#' resolve problemas de caracteres especiais.
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:///site.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# A SECRET_KEY é essencial para segurança de sessões e deve ser uma string longa e aleatória.
# Em produção, deve ser definida via variável de ambiente.
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "uma_chave_secreta_muito_segura_e_longa_para_producao")

db = SQLAlchemy(app)
bcrypt = Bcrypt(app) # Inicializa o Bcrypt para hash de senhas

# --- Definição dos Modelos (Tabelas) do Banco de Dados ---
# Cada classe representa uma tabela no banco de dados.

class User(db.Model):
    """
    Modelo para usuários do sistema (ex: administradores).
    Inclui um status para o fluxo de aprovação de cadastro.
    """
    __tablename__ = 'usuarios'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    status = db.Column(db.String(20), default='pending', nullable=False) # 'pending', 'active', 'inactive'
    acesso = db.Column(db.String(20), default='relator', nullable=False) # 'administrativo', 'relator'
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<User {self.username} - {self.status}>'

class Area(db.Model):
    """
    Modelo para as áreas de trabalho.
    Substitui a lista 'areas' do config.json.
    """
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), unique=True, nullable=False)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<Area {self.nome}>'

class Projeto(db.Model):
    """
    Modelo para os projetos.
    Substitui a lista 'projetos' do config.json.
    """
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), unique=True, nullable=False)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<Projeto {self.nome}>'

class Funcionario(db.Model):
    """
    Modelo para os funcionários.
    Substitui o employees.json. O 'id' é o identificador do QR Code.
    """
    id = db.Column(db.Integer, primary_key=True) # ID auto-gerado pelo PostgreSQL
    nome = db.Column(db.String(100), nullable=False)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    def __repr__(self):
        return f'<Funcionario {self.nome} ({self.id})>'

class Orcamento(db.Model):
    """
    Modelo para os orçamentos de horas por área/projeto/número.
    Substitui o orcamentos.json.
    """
    id = db.Column(db.Integer, primary_key=True)
    area_nome = db.Column(db.String(100), nullable=False)
    projeto_nome = db.Column(db.String(100), nullable=False)
    horas_orcadas = db.Column(db.Float, nullable=False)
    ativo = db.Column(db.Boolean, default=True, nullable=False)

    # Garante que a combinação de area e projeto seja única
    __table_args__ = (db.UniqueConstraint('area_nome', 'projeto_nome', name='_area_projeto_uc'),)

    def __repr__(self):
        return f'<Orcamento {self.area_nome}/{self.projeto_nome} - {self.horas_orcadas}h>'

class Registro(db.Model):
    """
    Modelo para os registros de horas trabalhadas.
    Substitui o armazenamento no Supabase e no Excel como fonte primária.
    """
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.Date, nullable=False)
    funcionario_id = db.Column(db.Integer, db.ForeignKey('funcionario.id'), nullable=False)
    funcionario_nome = db.Column(db.String(100), nullable=False)
    area_nome = db.Column(db.String(100), nullable=False)
    projeto_nome = db.Column(db.String(100), nullable=False)
    hora_inicio = db.Column(db.Time, nullable=False)
    hora_fim = db.Column(db.Time, nullable=True)
    status = db.Column(db.String(20), default='em_andamento')
    total_horas = db.Column(db.Float, nullable=True)

    # Relacionamento com Funcionario
    funcionario = db.relationship('Funcionario', backref=db.backref('registros', lazy=True))

    @property
    def duracao_formatada(self):
        """Calcula a duração total entre hora_inicio e hora_fim formatada (ex: 4h ou 4h 15min)."""
        if not self.hora_inicio or not self.hora_fim:
            return "-"
        
        dt_inicio = datetime.combine(date.min, self.hora_inicio)
        dt_fim = datetime.combine(date.min, self.hora_fim)
        
        # Lógica para registros que passam da meia-noite
        if dt_fim < dt_inicio:
            dt_fim += timedelta(days=1)
            
        delta = dt_fim - dt_inicio
        total_seconds = int(delta.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        
        if minutes == 0:
            return f"{hours}h"
        return f"{hours}h {minutes}min"

    def __repr__(self):
        return f'<Registro {self.data} - {self.funcionario_nome} ({self.hora_inicio}-{self.hora_fim})>'

class HistoricoAlteracaoHoras(db.Model):
    """
    Modelo para auditoria de alterações em registros de horas.
    """
    __tablename__ = 'historico_alteracoes_horas'
    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, db.ForeignKey('registro.id'), nullable=False)
    data_antiga = db.Column(db.Date)
    data_nova = db.Column(db.Date)
    hora_inicio_antiga = db.Column(db.Time)
    hora_inicio_nova = db.Column(db.Time)
    hora_fim_antiga = db.Column(db.Time)
    hora_fim_nova = db.Column(db.Time)
    motivo = db.Column(db.Text, nullable=False)
    area_antiga = db.Column(db.String(100))
    area_nova = db.Column(db.String(100))
    projeto_antigo = db.Column(db.String(100))
    projeto_novo = db.Column(db.String(100))
    criado_em = db.Column(db.DateTime, default=datetime.now)

# --- Variáveis Globais e Funções de Inicialização ---
# As variáveis globais AREAS, PROJETOS, CHARTS serão populadas do DB ou config.json.
AREAS = []
PROJETOS = []
CHARTS = ['bar'] # Default para tipos de gráficos

def load_initial_data_from_db():
    """
    Carrega as listas de áreas e projetos do banco de dados para as variáveis globais.
    Também tenta carregar a configuração de gráficos de um arquivo JSON.
    """
    global AREAS, PROJETOS, CHARTS
    with app.app_context():
        AREAS[:] = [area.nome for area in Area.query.filter_by(ativo=True).order_by(Area.nome).all()]
        PROJETOS[:] = [projeto.nome for projeto in Projeto.query.filter_by(ativo=True).order_by(Projeto.nome).all()]
        try:
            # Mantém a configuração de gráficos em config.json por enquanto
            with open('config.json', 'r', encoding='utf-8') as f:
                config_file = json.load(f)
                CHARTS[:] = config_file.get('charts', ['bar'])
        except FileNotFoundError:
            # Se config.json não existir, usa o default de CHARTS
            pass
        except json.JSONDecodeError:
            print("Erro ao ler config.json. Usando configuração de gráficos padrão.")
            pass

def _get_processed_report_data(areas=None, projetos=None, funcionarios=None, data_inicio=None, data_fim=None):
    """
    Calcula e retorna os dados processados para relatórios (horas trabalhadas, orçadas, etc.).
    Suporta filtros opcionais.
    Retorna:
        tuple: (all_registros, all_orcamentos_db, report_data)
            - all_registros: Lista de todos os objetos Registro.
            - all_orcamentos_db: Lista de todos os objetos Orcamento.
            - report_data: Lista de dicionários com dados processados para a aba 'Gráficos'.
    """
    query_reg = Registro.query
    query_orc = Orcamento.query

    if areas:
        query_reg = query_reg.filter(Registro.area_nome.in_(areas))
        query_orc = query_orc.filter(Orcamento.area_nome.in_(areas))
    if projetos:
        query_reg = query_reg.filter(Registro.projeto_nome.in_(projetos))
        query_orc = query_orc.filter(Orcamento.projeto_nome.in_(projetos))
    if funcionarios:
        query_reg = query_reg.filter(Registro.funcionario_id.in_(funcionarios))
    
    if data_inicio:
        if isinstance(data_inicio, str):
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        query_reg = query_reg.filter(Registro.data >= data_inicio)
    if data_fim:
        if isinstance(data_fim, str):
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        query_reg = query_reg.filter(Registro.data <= data_fim)

    all_registros = query_reg.order_by(Registro.data.desc(), Registro.hora_inicio.desc()).all() # Mantém histórico
    all_orcamentos_db = query_orc.filter_by(ativo=True).order_by(Orcamento.area_nome, Orcamento.projeto_nome).all()

    # Calcula horas trabalhadas por área/projeto a partir dos registros do DB
    horas_trabalhadas = {}
    for reg in all_registros:
        if reg.hora_inicio and reg.hora_fim and reg.area_nome and reg.projeto_nome:
            try:
                dt_inicio = datetime.combine(date.min, reg.hora_inicio)
                dt_fim = datetime.combine(date.min, reg.hora_fim)
                # Garante que a hora fim é maior que a hora início para cálculo correto
                if dt_fim < dt_inicio:
                    # Se a hora fim for no dia seguinte, adiciona 24 horas
                    dt_fim += timedelta(days=1)
                horas = (dt_fim - dt_inicio).total_seconds() / 3600
                key = f"{reg.area_nome} - {reg.projeto_nome}"
                horas_trabalhadas[key] = horas_trabalhadas.get(key, 0) + horas
            except Exception as e:
                print(f"Erro ao calcular horas para registro {reg.id}: {e}")
                pass

    report_data = []
    for orc in all_orcamentos_db:
        key = f"{orc.area_nome} - {orc.projeto_nome}"
        trabalhadas = horas_trabalhadas.get(key, 0)
        orcadas = orc.horas_orcadas
        restantes = max(0, orcadas - trabalhadas) # Horas restantes não podem ser negativas
        percentual = (trabalhadas / orcadas) * 100 if orcadas > 0 else 0

        report_data.append({
            "key": key,
            "area_nome": orc.area_nome,
            "projeto_nome": orc.projeto_nome,
            "horas_trabalhadas": round(trabalhadas, 2),
            "horas_orcadas": orcadas,
            "horas_restantes": round(restantes, 2),
            "percentual": round(percentual, 2)
        })
    
    return all_registros, all_orcamentos_db, report_data


# --- Funções de Geração de Excel e Gráficos (adaptadas para SQLAlchemy) ---
def criar_planilha_se_nao_existir():
    """
    Garante que o arquivo Excel exista ou seja criado.
    """
    if not os.path.exists(EXCEL_FILE):
        atualizar_graficos()

@app.route('/')
def index():
    """Renderiza a página inicial do leitor de QR Code."""
    criar_planilha_se_nao_existir() # Garante que o arquivo Excel existam
    return render_template('index.html')

@app.route('/verificar', methods=['POST'])
def verificar():
    """
    Verifica se há um registro de atividade "em andamento" para um funcionário.
    """
    data_req = request.get_json()
    idf = data_req.get("id")
    nome = data_req.get("nome")

    # Busca o último registro aberto do funcionário (status em_andamento)
    registro_aberto = Registro.query.filter_by(
        funcionario_id=idf,
        status='em_andamento'
    ).order_by(Registro.id.desc()).first()

    if registro_aberto:
        # Calcula tempo decorrido para exibição
        inicio_dt = datetime.combine(registro_aberto.data, registro_aberto.hora_inicio)
        decorrido = datetime.now() - inicio_dt
        segundos = int(decorrido.total_seconds())
        horas = segundos // 3600
        minutos = (segundos % 3600) // 60
        tempo_str = f"{horas} horas e {minutos} minutos" if horas > 0 else f"{minutos} minutos"

        return jsonify({
            "aberto": True,
            "area": registro_aberto.area_nome,
            "projeto": registro_aberto.projeto_nome,
            "inicio": registro_aberto.hora_inicio.strftime("%H:%M"),
            "data": registro_aberto.data.strftime("%d/%m/%Y"),
            "tempo_decorrido": tempo_str
        })
    else:
        return jsonify({"aberto": False, "area": "", "projeto": ""})

@app.route('/registrar', methods=['POST'])
def registrar():
    """
    Inicia ou Finaliza uma atividade automaticamente via QR Code.
    """
    data_payload = request.get_json()
    idf = data_payload.get("id")
    nome = data_payload.get("nome")
    tipo_acao = data_payload.get("tipoAcao") # 'iniciar' ou 'finalizar'

    agora = datetime.now()

    try:
        funcionario = Funcionario.query.get(idf)
        if not funcionario:
            funcionario = Funcionario(id=idf, nome=nome)
            db.session.add(funcionario)

        if tipo_acao == 'iniciar':
            area = data_payload.get("area")
            projeto = data_payload.get("projeto")
            
            novo_registro = Registro(
                data=agora.date(),
                funcionario_id=idf,
                funcionario_nome=nome,
                area_nome=area,
                projeto_nome=projeto,
                hora_inicio=agora.time(),
                status="em_andamento"
            )
            db.session.add(novo_registro)
            msg = "Atividade iniciada com sucesso."
        
        elif tipo_acao == 'finalizar':
            registro = Registro.query.filter_by(funcionario_id=idf, status='em_andamento').first()
            if not registro:
                return jsonify({"status": "error", "message": "Nenhum registro aberto encontrado."})
            
            registro.hora_fim = agora.time()
            registro.status = 'finalizado'
            
            # Calcula total de horas para o banco
            inicio_dt = datetime.combine(registro.data, registro.hora_inicio)
            diff = agora - inicio_dt
            registro.total_horas = diff.total_seconds() / 3600
            msg = "Atividade finalizada com sucesso."

        db.session.commit()

        # Após salvar no DB, atualiza o arquivo Excel
        atualizar_graficos()
        return jsonify({"status": "ok", "message": msg})
    except Exception as e:
        db.session.rollback() # Desfaz a transação em caso de erro
        print(f"Erro ao registrar no banco de dados: {e}")
        return jsonify({"status": "error", "message": f"Erro ao registrar: {e}"})

@app.route('/static/config.js')
def config_js():
    """
    Gera um arquivo JavaScript dinâmico com as configurações (áreas, projetos, gráficos).
    Agora carrega áreas e projetos do banco de dados.
    """
    # Carrega as áreas e projetos do banco de dados para o frontend
    areas_db = [area.nome for area in Area.query.filter_by(ativo=True).order_by(Area.nome).all()]
    projetos_db = [projeto.nome for projeto in Projeto.query.filter_by(ativo=True).order_by(Projeto.nome).all()]

    areas_json = json.dumps(areas_db)
    projetos_json = json.dumps(projetos_db)
    charts_json = json.dumps(CHARTS)
    js_code = f"""
    var AREAS = {areas_json};
    var PROJETOS = {projetos_json};
    var CHARTS = {charts_json};
    """
    return Response(js_code, mimetype='application/javascript')

@app.route('/api/employees')
def api_employees():
    """
    Retorna a lista de funcionários cadastrados no banco de dados.
    Substitui a leitura de employees.json.
    """
    employees = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    return jsonify([{"id": emp.id, "nome": emp.nome} for emp in employees])

@app.route('/login', methods=['POST'])
def login():
    """
    Autentica um usuário administrador.
    Verifica credenciais e o status de aprovação do usuário no banco de dados.
    """
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')

    user = User.query.filter_by(username=username, ativo=True).first()

    if user and bcrypt.check_password_hash(user.password_hash, password):
        if user.status == 'active':
            session['user_id'] = user.id
            session['username'] = user.username
            session['acesso'] = user.acesso
            return jsonify({"status": "ok"})
        else:
            return jsonify({"status": "error", "message": "FLASH: Conta aguardando validação ou inativa. Contate o administrador."})
    return jsonify({"status": "error", "message": "Credenciais inválidas"})

@app.route('/logout')
def logout():
    """Desloga o usuário da sessão."""
    session.pop('user_id', None)
    session.pop('username', None)
    session.pop('acesso', None)
    return redirect(url_for('index'))

@app.route('/admin')
def admin():
    """
    Renderiza o painel de administração.
    Requer autenticação e carrega dados do banco de dados.
    """
    if 'user_id' not in session:
        # Redireciona para a página inicial se não estiver logado
        return redirect(url_for('index'))
    
    if session.get('acesso') != 'administrativo':
        # Se for relator, redireciona para a página de relatórios
        return redirect(url_for('view_reports'))

    # Processamento de Filtros
    f_areas = request.args.getlist('f_areas')
    f_projetos = request.args.getlist('f_projetos')
    f_funcionarios = request.args.getlist('f_funcionarios')
    f_date_preset = request.args.get('f_date_preset', 'all')
    f_inicio = request.args.get('f_inicio')
    f_fim = request.args.get('f_fim')

    # Lógica de presets de data
    if f_date_preset != 'custom' and f_date_preset != 'all':
        hoje = date.today()
        if f_date_preset == 'today':
            f_inicio = f_fim = hoje.strftime('%Y-%m-%d')
        elif f_date_preset == 'yesterday':
            ontem = hoje - timedelta(days=1)
            f_inicio = f_fim = ontem.strftime('%Y-%m-%d')
        elif f_date_preset == '7days':
            f_inicio = (hoje - timedelta(days=6)).strftime('%Y-%m-%d')
            f_fim = hoje.strftime('%Y-%m-%d')
        elif f_date_preset == '30days':
            f_inicio = (hoje - timedelta(days=29)).strftime('%Y-%m-%d')
            f_fim = hoje.strftime('%Y-%m-%d')
        elif f_date_preset == 'month':
            f_inicio = hoje.replace(day=1).strftime('%Y-%m-%d')
            f_fim = hoje.strftime('%Y-%m-%d')
    
    # Determina qual view abrir por padrão
    default_view = request.args.get('view', 'home')

    all_registros, all_orcamentos_db, report_data = _get_processed_report_data(
        areas=f_areas,
        projetos=f_projetos,
        funcionarios=f_funcionarios,
        data_inicio=f_inicio if f_inicio else None,
        data_fim=f_fim if f_fim else None
    )

    # Calcula o total de horas dos registros filtrados
    total_horas_filtradas = sum(reg.total_horas for reg in all_registros if reg.total_horas)

    # Carrega todos os dados necessários para exibir no painel de administração
    areas = Area.query.filter_by(ativo=True).order_by(Area.nome).all()
    projetos = Projeto.query.filter_by(ativo=True).order_by(Projeto.nome).all()
    employees = Funcionario.query.filter_by(ativo=True).order_by(Funcionario.nome).all()
    users = User.query.filter_by(ativo=True).order_by(User.username).all() # Para gestão de usuários

    return render_template('admin.html',
                           areas=areas,
                           projetos=projetos,
                           employees=employees,
                           orcamentos=all_orcamentos_db,
                           users=users,
                           registros=all_registros,
                           total_horas_filtradas=total_horas_filtradas,
                           f_areas=f_areas,
                           f_projetos=f_projetos,
                           f_funcionarios=f_funcionarios,
                           f_date_preset=f_date_preset,
                           f_inicio=f_inicio,
                           f_fim=f_fim,
                           default_view=default_view)

# --- Rotas de Administração (CRUD para o Banco de Dados) ---

@app.route('/admin/add_area', methods=['POST'])
def add_area():
    """Adiciona uma nova área ao banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    area_nome = data.get('area')
    if area_nome:
        existing_area = Area.query.filter_by(nome=area_nome).first()
        if existing_area:
            if not existing_area.ativo:
                existing_area.ativo = True # Reativa se já existia inativa
                db.session.commit()
                load_initial_data_from_db()
                return jsonify({"status": "ok"})
            return jsonify({"status": "error", "message": "Área já existe"})
        else:
            new_area = Area(nome=area_nome)
            db.session.add(new_area)
        db.session.commit()
        load_initial_data_from_db()
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Nome da área inválido"})

@app.route('/admin/delete_area', methods=['POST'])
def delete_area():
    """Exclui uma área do banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    area_nome = data.get('area')
    area = Area.query.filter_by(nome=area_nome).first()
    if area:
        area.ativo = False # Soft Delete
        db.session.commit()
        load_initial_data_from_db()
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Área não encontrada"})

@app.route('/admin/add_projeto', methods=['POST'])
def add_projeto():
    """Adiciona um novo projeto ao banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    projeto_nome = data.get('projeto')
    if projeto_nome:
        existing_projeto = Projeto.query.filter_by(nome=projeto_nome).first()
        if existing_projeto:
            if not existing_projeto.ativo:
                existing_projeto.ativo = True # Reativa se já existia inativa
                db.session.commit()
                load_initial_data_from_db()
                return jsonify({"status": "ok"})
            return jsonify({"status": "error", "message": "Projeto já existe"})
        else:
            new_projeto = Projeto(nome=projeto_nome)
            db.session.add(new_projeto)
        db.session.commit()
        load_initial_data_from_db()
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Nome do projeto inválido"})

@app.route('/admin/delete_projeto', methods=['POST'])
def delete_projeto():
    """Exclui um projeto do banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    projeto_nome = data.get('projeto')
    projeto = Projeto.query.filter_by(nome=projeto_nome).first()
    if projeto:
        projeto.ativo = False # Soft Delete
        db.session.commit()
        load_initial_data_from_db()
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Projeto não encontrado"})

@app.route('/admin/delete_employee', methods=['POST'])
def delete_employee():
    """Exclui um funcionário do banco de dados após validações."""
    if session.get('acesso') != 'administrativo': 
        return jsonify({"status": "error", "message": "Não autorizado"})
    
    data = request.get_json()
    idf = data.get('id')
    
    employee = Funcionario.query.get(idf)
    if not employee:
        return jsonify({"status": "error", "message": "Funcionário não encontrado."})
    
    # Regra de Negócio: Verificar se há atividade em andamento
    registro_aberto = Registro.query.filter_by(funcionario_id=idf, status='em_andamento').first()
    if registro_aberto:
        return jsonify({"status": "error", "message": "Não é possível excluir um funcionário com atividade em andamento."})
    
    try:
        # Soft Delete do funcionário, mantendo registros históricos
        employee.ativo = False
        db.session.commit()
        return jsonify({"status": "ok"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Erro ao excluir: {e}"})

@app.route('/admin/add_employee', methods=['POST'])
def add_employee():
    """Adiciona um novo funcionário ao banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    nome = data.get('nome')

    if not nome:
        return jsonify({"status": "error", "message": "Informe o nome do funcionário."})

    try:
        new_employee = Funcionario(nome=nome, ativo=True)
        db.session.add(new_employee)
        db.session.commit()
        return jsonify({"status": "ok"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Erro ao salvar: {e}"})

@app.route('/admin/edit_registro', methods=['POST'])
def edit_registro():
    """Edita um registro de horas e salva no histórico de auditoria."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    
    data_req = request.get_json()
    reg_id = data_req.get('id')
    nova_data_str = data_req.get('data')
    nova_inicio_str = data_req.get('hora_inicio')
    nova_fim_str = data_req.get('hora_fim')
    nova_area_nome = data_req.get('area')
    novo_projeto_nome = data_req.get('projeto')
    motivo = data_req.get('motivo')

    if not motivo:
        return jsonify({"status": "error", "message": "O motivo da alteração é obrigatório."})

    registro = Registro.query.get(reg_id)
    if not registro:
        return jsonify({"status": "error", "message": "Registro não encontrado."})

    # Validações para Área e Projeto
    if nova_area_nome:
        area_existe = Area.query.filter_by(nome=nova_area_nome, ativo=True).first()
        if not area_existe:
            return jsonify({"status": "error", "message": f"Área '{nova_area_nome}' não encontrada ou inativa."})
    if novo_projeto_nome:
        projeto_existe = Projeto.query.filter_by(nome=novo_projeto_nome, ativo=True).first()
        if not projeto_existe:
            return jsonify({"status": "error", "message": f"Projeto '{novo_projeto_nome}' não encontrado ou inativo."})

    try:
        # Criar log de auditoria
        historico = HistoricoAlteracaoHoras(
            registro_id=registro.id,
            data_antiga=registro.data,
            area_antiga=registro.area_nome,
            projeto_antigo=registro.projeto_nome,
            hora_inicio_antiga=registro.hora_inicio,
            hora_fim_antiga=registro.hora_fim,
            motivo=motivo
        )

        # Atualizar registro
        if nova_data_str:
            registro.data = datetime.strptime(nova_data_str, '%Y-%m-%d').date()
        if nova_inicio_str:
            registro.hora_inicio = datetime.strptime(nova_inicio_str, '%H:%M').time()
        if nova_fim_str:
            registro.hora_fim = datetime.strptime(nova_fim_str, '%H:%M').time()
        if nova_area_nome:
            registro.area_nome = nova_area_nome
        if novo_projeto_nome:
            registro.projeto_nome = novo_projeto_nome

        # Recalcular total de horas
        dt_inicio = datetime.combine(registro.data, registro.hora_inicio)
        dt_fim = datetime.combine(registro.data, registro.hora_fim)
        if dt_fim < dt_inicio:
            dt_fim += timedelta(days=1)
        registro.total_horas = (dt_fim - dt_inicio).total_seconds() / 3600

        # Completar log de auditoria
        historico.data_nova = registro.data
        historico.area_nova = registro.area_nome
        historico.projeto_nova = registro.projeto_nome
        historico.hora_inicio_nova = registro.hora_inicio
        historico.hora_fim_nova = registro.hora_fim

        db.session.add(historico)
        db.session.commit()
        atualizar_graficos()
        return jsonify({"status": "ok", "message": "Registro atualizado com sucesso"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Erro ao editar: {e}"})

@app.route('/admin/add_orcamento', methods=['POST'])
def add_orcamento():
    """Adiciona ou atualiza um orçamento de horas no banco de dados."""
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    area = data.get('area')
    projeto = data.get('projeto')
    horasOrcadas = data.get('horasOrcadas')
    if not all([area, projeto, horasOrcadas is not None]):
        return jsonify({"status": "error", "message": "Dados de orçamento inválidos"})

    try:
        horasOrcadas = float(horasOrcadas)
    except ValueError:
        return jsonify({"status": "error", "message": "Horas orçadas devem ser um número válido."})

    # Tenta encontrar um orçamento existente para a combinação de área/projeto/número
    orcamento = Orcamento.query.filter_by(
        area_nome=area,
        projeto_nome=projeto
    ).first()

    if orcamento:
        # Se existir, atualiza as horas orçadas
        orcamento.horas_orcadas = horasOrcadas
        orcamento.ativo = True # Garante que está ativo ao atualizar
    else:
        # Se não existir, cria um novo orçamento
        orcamento = Orcamento(
            area_nome=area,
            projeto_nome=projeto,
            horas_orcadas=horasOrcadas
        )
        db.session.add(orcamento)
    db.session.commit()
    atualizar_graficos() # Atualiza os gráficos após a alteração do orçamento
    return jsonify({"status": "ok"})

@app.route('/admin/select_charts', methods=['POST'])
def select_charts():
    """
    Permite ao administrador selecionar quais tipos de gráficos serão gerados.
    A configuração é salva em config.json.
    """
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    charts = data.get('charts', [])
    global CHARTS
    CHARTS[:] = charts

    # Persiste a configuração de gráficos em config.json
    try:
        config_file = {}
        if os.path.exists('config.json'):
            with open('config.json', 'r', encoding='utf-8') as f:
                config_file = json.load(f)
        config_file['charts'] = CHARTS
        with open('config.json', 'w', encoding='utf-8') as f:
            json.dump(config_file, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Erro ao salvar configuração de gráficos em config.json: {e}")
    return jsonify({"status": "ok"})

@app.route('/api/orcamentos')
def api_orcamentos():
    """
    Retorna a lista de orçamentos cadastrados no banco de dados.
    Substitui a leitura de orcamentos.json.
    """
    orcamentos = Orcamento.query.filter_by(ativo=True).order_by(Orcamento.area_nome, Orcamento.projeto_nome).all()
    return jsonify([
        {
            "area": o.area_nome,
            "projeto": o.projeto_nome,
            "horasOrcadas": o.horas_orcadas
        } for o in orcamentos
    ])

@app.route('/qrcodes')
def qrcodes():
    return render_template('qrcodes.html')

# --- Rotas para Gestão de Usuários (Admin) ---

@app.route('/register_user', methods=['POST'])
def register_user():
    """
    Permite que novos usuários se registrem no sistema.
    O status inicial é 'pending' e requer aprovação.
    """
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    email = data.get('email')

    if not username or not password or not email:
        return jsonify({"status": "error", "message": "Usuário, senha e e-mail são obrigatórios."})

    existing_user = User.query.filter((User.username == username) | (User.email == email)).first()
    if existing_user:
        return jsonify({"status": "error", "message": "Usuário ou e-mail já cadastrado."})

    # Gera o hash da senha antes de salvar no banco de dados
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    new_user = User(username=username, password_hash=hashed_password, email=email, status='pending')
    db.session.add(new_user)
    db.session.commit()
    return jsonify({"status": "ok", "message": "Usuário registrado com sucesso. Aguardando aprovação do administrador."})

@app.route('/admin/update_user_status', methods=['POST'])
def update_user_status():
    """
    Permite que um administrador altere o status de um usuário (pending, active, inactive).
    """
    if session.get('acesso') != 'administrativo': return jsonify({"status": "error", "message": "Não autorizado"})
    data = request.get_json()
    user_id = data.get('user_id')
    new_status = data.get('status')

    user = User.query.get(user_id)
    if user:
        user.status = new_status
        db.session.commit()
        return jsonify({"status": "ok"})
    return jsonify({"status": "error", "message": "Usuário não encontrado."})

@app.route('/admin/download_excel')
def download_excel():
    """Permite ao administrador baixar o arquivo Excel de registros."""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    # Atualiza os gráficos/dados antes de enviar para garantir que o arquivo esteja em dia
    atualizar_graficos()
    
    if os.path.exists(EXCEL_FILE):
        return send_file(os.path.abspath(EXCEL_FILE), as_attachment=True)
    return "Arquivo de registros ainda não foi gerado.", 404

@app.route('/admin/reports')
def view_reports():
    """Exibe os relatórios de horas trabalhadas e orçadas diretamente no navegador."""
    if 'user_id' not in session:
        return redirect(url_for('index'))
    
    if session.get('acesso') != 'relator':
        # Se for administrativo, redireciona para o painel principal
        return redirect(url_for('admin'))

    # Filtros do Relator
    f_areas = request.args.getlist('f_areas')
    f_projetos = request.args.getlist('f_projetos')

    all_registros, all_orcamentos_db, report_data = _get_processed_report_data(
        areas=f_areas,
        projetos=f_projetos
    )
    
    # Carrega opções para os filtros
    areas = Area.query.filter_by(ativo=True).order_by(Area.nome).all()
    projetos = Projeto.query.filter_by(ativo=True).order_by(Projeto.nome).all()

    return render_template('reports.html', 
                           report_data=report_data, # Dados processados para a tabela de gráficos
                           areas=areas,
                           projetos=projetos,
                           f_areas=f_areas,
                           f_projetos=f_projetos)

def atualizar_graficos():
    """
    Gera um novo arquivo Excel 'registros.xlsx' do zero para evitar corrupção.
    Popula as abas 'Registros', 'Orçamentos' e 'Gráficos' com dados do banco.
    """
    all_registros, all_orcamentos_db, report_data = _get_processed_report_data()
    
    wb = Workbook()

    # --- Aba Registros ---
    ws_reg = wb.active
    ws_reg.title = "Registros"
    headers_reg = ["Data", "ID", "Nome", "Área", "Projeto", "Hora Início", "Hora Fim", "Hora Total"]
    ws_reg.append(headers_reg)
    
    # Estilo dos cabeçalhos
    for cell in ws_reg[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Largura das colunas
    column_widths_reg = [12, 10, 30, 20, 15, 15, 15, 15]
    for i, width in enumerate(column_widths_reg, 1):
        ws_reg.column_dimensions[ws_reg.cell(row=1, column=i).column_letter].width = width

    for reg in all_registros:
        ws_reg.append([
            reg.data.strftime("%Y-%m-%d"),
            reg.funcionario_id,
            reg.funcionario_nome,
            reg.area_nome,
            reg.projeto_nome,
            reg.hora_inicio.strftime("%H:%M"),
            reg.hora_fim.strftime("%H:%M") if reg.hora_fim else "Em aberto",
            reg.duracao_formatada
        ])

    # --- Aba Orçamentos ---
    ws_orc = wb.create_sheet("Orçamentos")
    headers_orc = ["Área", "Projeto", "Horas Orçadas"]
    ws_orc.append(headers_orc)
    
    for cell in ws_orc[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    column_widths_orc = [30, 30, 18]
    for i, width in enumerate(column_widths_orc, 1):
        ws_orc.column_dimensions[ws_orc.cell(row=1, column=i).column_letter].width = width

    for orc in all_orcamentos_db:
        ws_orc.append([
            orc.area_nome,
            orc.projeto_nome,
            orc.horas_orcadas
        ])

    # --- Aba Gráficos ---
    ws_chart = wb.create_sheet("Gráficos")
    headers_chart = ["Área/Projeto", "Horas Trabalhadas", "Horas Orçadas", "Horas Restantes", "Percentual (%)"]
    ws_chart.append(headers_chart)
    
    for cell in ws_chart[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    column_widths_chart = [30, 18, 18, 18, 18]
    for i, width in enumerate(column_widths_chart, 1):
        ws_chart.column_dimensions[ws_chart.cell(row=1, column=i).column_letter].width = width

    row = 2
    for data_item in report_data:
        ws_chart.cell(row=row, column=1).value = data_item["key"]
        ws_chart.cell(row=row, column=2).value = data_item["horas_trabalhadas"]
        ws_chart.cell(row=row, column=3).value = data_item["horas_orcadas"]
        ws_chart.cell(row=row, column=4).value = f"=C{row}-B{row}"
        ws_chart.cell(row=row, column=5).value = f"=IF(C{row}>0, (B{row}/C{row})*100, 0)"
        
        ws_chart.cell(row=row, column=2).number_format = '0.00'
        ws_chart.cell(row=row, column=3).number_format = '0'
        ws_chart.cell(row=row, column=4).number_format = '0.00'
        ws_chart.cell(row=row, column=5).number_format = '0.00"%"'
        row += 1

    # Tabela e Gráficos
    if row > 2:
        tab_ref = f"A1:E{row-1}"
        table = Table(displayName="DadosGraficos", ref=tab_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=True)
        ws_chart.add_table(table)

        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=row-1)
        if 'bar' in CHARTS:
            bar_chart = BarChart()
            bar_chart.add_data(Reference(ws_chart, min_col=2, min_row=1, max_col=4, max_row=row-1), titles_from_data=True)
            bar_chart.set_categories(cats)
            ws_chart.add_chart(bar_chart, "F2")

        if 'doughnut' in CHARTS:
            doughnut_chart = DoughnutChart()
            doughnut_chart.add_data(Reference(ws_chart, min_col=5, min_row=2, max_row=row-1), titles_from_data=False)
            doughnut_chart.set_categories(cats)
            ws_chart.add_chart(doughnut_chart, "F20")

    wb.save(EXCEL_FILE)
    print(f"Arquivo Excel '{EXCEL_FILE}' gerado com sucesso.")

# --- Inicialização da Aplicação ---

# Esta função garante que o banco de dados e as tabelas existam
# independentemente de como o app é iniciado (Gunicorn ou Python direto)
with app.app_context():
    db.create_all()
    # Garante que o arquivo Excel e as abas existam
    criar_planilha_se_nao_existir()
    load_initial_data_from_db()

if __name__ == '__main__':
    with app.app_context():
        # Cria um usuário administrador padrão se não houver nenhum.
        # Isso é útil para o primeiro acesso ao painel de administração.
        if User.query.filter_by(username='admin').first() is None:
            hashed_password = bcrypt.generate_password_hash('admin').decode('utf-8')
            admin_user = User(username='admin', password_hash=hashed_password, email='admin@fleximedical.com.br', status='active', acesso='administrativo')
            db.session.add(admin_user)
            db.session.commit()
            print("Usuário 'admin' criado com senha 'admin' e status 'active'.")

    port = int(os.environ.get('PORT', 5000))
    # Em ambiente de produção, debug=False e host='0.0.0.0' são recomendados.
    app.run(host='0.0.0.0', port=port, debug=True)
